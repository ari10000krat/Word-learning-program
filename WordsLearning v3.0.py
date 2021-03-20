from tkinter import *
import tkinter
from tkinter import filedialog
import PIL
import os
import openpyxl
import random
from copy import deepcopy
from tkinter import messagebox
from PIL import ImageTk, Image



def xlsxOpen():
    """
    Открывает файл EXEL
    :return: моссив 1 колонка DE 2 колонка RUS
    """
    global wordsArrIsReady
    global wordsArr
    global root
    root.quit()
    file_name = None
    while file_name == None:
        file_name = filedialog.askopenfile(filetypes=(('EXEL files', '*.xlsx'),
                                                      ('All files', '*.*')))
    wb = openpyxl.load_workbook(file_name.name)
    sheet = wb.active
    wordsArr = []
    for i in range(0, sheet.max_row):
        wordsArr.append([sheet.cell(row=i + 1, column=1).value, sheet.cell(row=i + 1, column=2).value])
    wordsArrIsReady = True
    welcomeWindow()


def welcomeWindow():
    def exitButton():
        sys.exit()

    global root
    global mode
    global wordsArrIsReady
    # Настройки формы########################
    imageFile = PIL.Image.open('flagi3.png')
    img = ImageTk.PhotoImage(imageFile)
    root.geometry('750x400')
    root.config(bg=lightBlue)
    root.title = 'WelcomeWindow'
    root.resizable(False, False)

    def switchMode():
        global mode
        if mode == 'DE -> RU':
            mode = 'RU -> DE'
        else:
            mode = 'DE -> RU'
        LabelMode.config(text=f'Mode: {mode}')

    # Настройки приветствия########################
    LabelMode = Label(text=f'Mode: {mode}',
                      fg='#4888FC',
                      bg=lightBlue,
                      font=("Times", "30", "bold italic"))
    LabelMode.place(x=5, y=5)

    # Настройка картинки##########
    ImageLabel = Label(image=img)
    ImageLabel.place(x=5, y=150)

    # Настройки Меню################
    MenuFrame = Frame(bg='#AECAFC')

    LabelMenu = Label(MenuFrame,
                      text='Меню',
                      activebackground='#7CA9FC',
                      fg='#4888FC',
                      bg=lightBlue,
                      font=("Times", "30", "bold italic"))
    if wordsArrIsReady:
        button1 = Button(MenuFrame,
                         text='1. Учить слова',
                         activebackground='#7CA9FC',
                         fg='#4888FC',
                         bg=lightBlue,
                         width=20,
                         relief='groove',
                         font=myFont,
                         command=learningWindow)
    else:
        button1 = Button(MenuFrame,
                         text='1. Учить слова',
                         activebackground='#7CA9FC',
                         fg='#4888FC',
                         bg='red',
                         width=20,
                         relief='groove',
                         font=myFont)

    button3 = Button(MenuFrame,
                     text='2. Открыть словарь',
                     activebackground='#7CA9FC',
                     fg='#4888FC',
                     bg=lightBlue,
                     width=20,
                     relief='groove',
                     font=myFont,
                     command=xlsxOpen)

    button5 = Button(MenuFrame,
                     text='3. Сменить режим',
                     activebackground='#7CA9FC',
                     fg='#4888FC',
                     bg=lightBlue,
                     width=20,
                     relief='groove',
                     font=myFont,
                     command=switchMode)

    button6 = Button(MenuFrame,
                     text='4. Выход',
                     activebackground='#7CA9FC',
                     fg='#4888FC',
                     bg=lightBlue,
                     width=20,
                     relief='groove',
                     font=myFont,
                     command=exitButton)

    LabelMenu.pack()
    button1.pack()
    # button2.pack()
    button3.pack()
    # button4.pack()
    button5.pack()
    button6.pack()
    MenuFrame.place(x=400, y=100)

    root.mainloop()


def learningWindow():
    def SaveWindow():
        def SaveToExel():
            nonlocal current_arr
            nonlocal swWindow
            wb = openpyxl.Workbook()
            sheet = wb.active
            for i in range(len(current_arr)):
                for j in range(2):
                    sheet.cell(row=i + 1, column=j + 1).value = current_arr[i][j]
            wb.save(f'{nmEntry.get()}.xlsx')
            swWindow.destroy()

        # nonlocal CurrentArr
        swWindow = Tk()
        swWindow['bg'] = lightBlue
        swLabel = Label(swWindow,
                        text='NAME:',
                        font=myFont,
                        bg=lightBlue)
        swLabel.pack()
        nmEntry = Entry(swWindow,
                        font=myFont,
                        bg=lightBlue)
        nmEntry.pack()
        okButton = Button(swWindow,
                          text='SAVE',
                          bg=lightBlue,
                          font=myFont,
                          command=SaveToExel)
        okButton.pack()

    def StartPenaltyRound():
        nonlocal current_arr
        nonlocal word_index
        nonlocal rounds
        nonlocal fail_arr
        nonlocal count_of_words
        nonlocal count_of_true

        current_arr = fail_arr
        count_of_words = len(current_arr)
        count_of_true = 0  # НОВОЕ
        fail_arr = []
        random.shuffle(current_arr)
        word_index = 0
        rounds += 1

    def StartNewRound():
        nonlocal rounds
        nonlocal progress
        nonlocal count_of_words
        nonlocal current_arr
        nonlocal fail_arr
        nonlocal word_index
        nonlocal count_of_true
        rounds += 1
        progress = 0
        count_of_words = len(wordsArr)
        current_arr = deepcopy(wordsArr)
        random.shuffle(current_arr)
        fail_arr = []
        word_index = 0
        count_of_true = 0

    def UpdateStatistics():
        nonlocal progress
        nonlocal window1
        progress = int(count_of_true / count_of_words * 100)
        statisticsLabel.config(
            text=f'Round: {rounds}\nWords left: {count_of_words - count_of_true}\nProgress: {progress}%')

    def UpdateWords():
        nonlocal window1
        wordLabel.config(text=current_arr[word_index][mainIndex])
        wordLabelHelp.config(text=current_arr[word_index][secondIndex])

    def IsComplieted():
        nonlocal count_of_true
        nonlocal current_arr
        nonlocal word_index
        if word_index == len(current_arr) - 1:  # конец массива
            if count_of_true != len(current_arr) - 1 and len(current_arr) != 1:
                return -1
            else:  # ошибок нет
                return 1
        else:  # не конец списка
            return 0

    def yesClick():
        nonlocal count_of_true
        nonlocal word_index
        nonlocal window1
        nonlocal fail_arr
        nonlocal count_of_words
        nonlocal current_arr
        nonlocal progress
        nonlocal rounds
        global wordsArr
        m = IsComplieted()
        if m == 0:  # Есть еще слова
            count_of_true += 1
            word_index += 1
            UpdateStatistics()
            UpdateWords()
        elif m == -1:  # Есть ошибки
            StartPenaltyRound()
            UpdateStatistics()
            UpdateWords()
        elif m == 1:
            UpdateStatistics()
            messagebox.showinfo('ИНФА', 'Вы выучили слова')
            StartNewRound()
            UpdateWords()
            UpdateStatistics()

    def noClick():
        nonlocal count_of_true
        nonlocal word_index
        nonlocal window1
        nonlocal fail_arr
        nonlocal current_arr
        nonlocal count_of_words
        m = IsComplieted()
        if m == 0:
            if current_arr[word_index] not in fail_arr:
                fail_arr.append(current_arr[word_index])
            word_index += 1
            UpdateStatistics()
            UpdateWords()
        elif m == -1:
            if current_arr[word_index] not in fail_arr:
                fail_arr.append(current_arr[word_index])
            StartPenaltyRound()
            UpdateStatistics()
            UpdateWords()

    def helpClick():
        if wordLabelHelp['fg'] == lightBlue:
            wordLabelHelp['fg'] = 'black'
        else:
            wordLabelHelp['fg'] = lightBlue


    # Настройка формы***************************************************************************************
    global wordsArr
    global mode
    rounds = -1
    progress = 0
    count_of_words = 0
    current_arr = []
    fail_arr = []
    word_index = 0
    count_of_true = 0

    if mode == 'DE -> RU':
        mainIndex = 0
        secondIndex = 1
    elif mode == 'RU -> DE':
        mainIndex = 1
        secondIndex = 0

    StartNewRound()

    window1 = Tk()
    window1.geometry('800x300')
    window1.config(bg='#AECAFC')
    window1.resizable(False, False)



    # Настройка вывода слова
    LabelFrame = Frame(
        window1,
        bg=lightBlue
    )

    wordLabel = Label(LabelFrame,
                      text=current_arr[word_index][mainIndex],
                      bg=lightBlue,
                      font=("Times", "24", "bold"),
                      justify='center',
                      width=43)
    wordLabel.pack()

    wordLabelHelp = Label(LabelFrame,
                          text=current_arr[word_index][secondIndex],
                          fg=lightBlue,
                          bg=lightBlue,
                          font=("Times", "24", "bold"),
                          justify='center',
                          width=43)
    wordLabelHelp.pack()

    LabelFrame.place(x=5, y=5)
    buttonsFrame = Frame(window1,
                         bg='gray'
                         )

    buttonYes = Button(buttonsFrame,
                       text='YES',
                       bg='green2',
                       font=("Times", "24"),
                       justify='center',
                       width=10,
                       command=yesClick
                       )
    buttonYes.pack(side=LEFT)

    buttonNo = Button(buttonsFrame,
                      text='NO',
                      bg='red3',
                      font=("Times", "24"),
                      justify='center',
                      width=10,
                      command=noClick
                      )
    buttonNo.pack(side=LEFT)

    buttonHelp = Button(buttonsFrame,
                        text='HELP',
                        bg='orange2',
                        font=("Times", "24"),
                        justify='center',
                        width=10,
                        command=helpClick
                        )
    buttonHelp.pack(side=LEFT)

    buttonsFrame.place(x=5, y=140)

    buttonSave = Button(window1,
                        text='SAVE',
                        bg='cyan3',
                        font=("Times", "24"),
                        justify='center',
                        width=31,
                        command=SaveWindow)
    buttonSave.place(x=5, y=202)

    statisticsLabel = Label(window1,
                            text=f'Round: {rounds}\nWords left: {count_of_words - count_of_true}\nProgress: {progress}%',
                            font=myFont,
                            bg=lightBlue,
                            justify='left',
                            width=15)
    statisticsLabel.place(x=580, y=140)

    window1.mainloop()




if __name__ == '__main__':
    # Переменные####################################################################
    myFont = ('Times', '20')
    root = Tk()
    mode = 'RU -> DE'
    lightBlue = '#AECAFC'
    wordsArrIsReady = False
    wordsArr = []
    welcomeWindow()